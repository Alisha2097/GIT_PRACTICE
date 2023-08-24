from openpyxl import Workbook
wb = Workbook()

ws =wb.active

# ws['A1']="Hello Test"
# ws['C1']="Hello Test"

# testdata = [['Name','City'],['Ramesh','Kathmandu'],['Ram','Pokahra'],['Sita'],['Butwal']]
# for data in testdata:
#     ws.append(data)

for i in range (1,6):
    for j in range(1,5):
        ws.cell(row=i, column=j).value =i+j


wb.save("demoexcel.xlsx")