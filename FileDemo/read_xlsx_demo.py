from openpyxl import Workbook, load_workbook

wb = load_workbook(filename='demoexcel.xlsx')
sh = wb.active 
# sh = wb['sheet_name'] to read the data from the particular excel sheet 

# print(sh['A1'].value)
# print(wb['sheet_name']['A2'].value)
# print(sh.cell(2,3).value)
# print(sh.cell(row=2,column=3).value) #preferred way 

row_ct = sh.max_row
col_ct = sh.max_column

for i in range (1,row_ct+1):
    for j in range(1,col_ct+1):
        print(sh.cell(row=i, column=j).value, end='')
        print('\n')

print(row_ct)
print(col_ct)